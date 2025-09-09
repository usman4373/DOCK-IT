import os
import subprocess
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re

def run_plip_analysis(complexes_dir, output_dir):
    """
    Run PLIP analysis on all complex files in the complexes directory
    and generate interaction reports in both TXT and XLSX formats.
    """
    # Create plip_processed_files directory
    plip_output_dir = os.path.join(output_dir, "plip_processed_files")
    os.makedirs(plip_output_dir, exist_ok=True)
    
    # Get all complex files
    complex_files = glob.glob(os.path.join(complexes_dir, "*.pdb"))
    
    if not complex_files:
        return {}
    
    # Run PLIP command
    cmd = ["plip", "-f"] + complex_files + ["-o", plip_output_dir, "-t"]
    
    try:
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode != 0:
            error_msg = f"PLIP error: {result.stderr}"
            log_error(plip_output_dir, "PLIP command", error_msg)
            return {}
    except Exception as e:
        error_msg = f"PLIP execution failed: {str(e)}"
        log_error(plip_output_dir, "PLIP execution", error_msg)
        return {}
    
    # Process PLIP results
    interaction_data = process_plip_results(plip_output_dir, complexes_dir)
    
    return interaction_data

def process_plip_results(plip_output_dir, complexes_dir):
    """
    Process PLIP results by converting TXT files to XLSX and extracting interaction data.
    """
    interaction_data = {}
    error_log = []
    
    # Find all report.txt files in subdirectories
    report_files = glob.glob(os.path.join(plip_output_dir, "**", "report.txt"), recursive=True)
    
    for report_file in report_files:
        try:
            # Get the complex name from the directory name
            dir_name = os.path.basename(os.path.dirname(report_file))
            complex_name = dir_name  # PLIP uses the complex filename as directory name
            
            # Convert TXT to XLSX and move to complexes directory
            xlsx_path = os.path.join(complexes_dir, f"{complex_name}.xlsx")
            convert_plip_txt_to_xlsx(report_file, xlsx_path)
            
            # Get interaction data from the XLSX file
            num_interactions, interaction_types = get_interaction_data_from_xlsx(xlsx_path)
            
            # Store interaction data
            interaction_data[complex_name] = {
                "num_interactions": num_interactions,
                "interaction_types": interaction_types
            }
            
        except Exception as e:
            error_msg = f"Error processing {report_file}: {str(e)}"
            error_log.append(error_msg)
    
    # Log any errors
    if error_log:
        log_error(plip_output_dir, "PLIP processing", "\n".join(error_log))
    
    return interaction_data

def parse_plip_report(file_path):
    """
    Parse a PLIP report file to extract interaction counts and types.
    """
    interaction_types = set()
    num_interactions = 0
    
    # Mapping from PLIP section names to standardized names
    interaction_mapping = {
        "HYDROPHOBIC INTERACTIONS": "Hydrophobic Interactions",
        "HYDROGEN BONDS": "Hydrogen Bonds",
        "WATER-BRIDGED HYDROGEN BONDS": "Water Bridges",
        "PI-STACKING": "π-Stacking",
        "PI-CATION INTERACTIONS": "π-Cation Interactions",
        "HALOGEN BONDS": "Halogen Bonds",
        "SALT BRIDGES": "Salt Bridges",
        "METAL COMPLEXES": "Metal Complexation"
    }
    
    with open(file_path, 'r') as f:
        content = f.read()
    
    # Split content into sections
    sections = re.split(r'\n\s*\*\*', content)
    
    for section in sections:
        # Check if this is an interaction section
        for plip_name, std_name in interaction_mapping.items():
            if plip_name in section:
                # Count the number of interactions in this section
                lines = section.split('\n')
                interaction_count = sum(1 for line in lines if '|' in line and not line.startswith('|') and not line.endswith('|'))
                
                if interaction_count > 0:
                    interaction_types.add(std_name)
                    num_interactions += interaction_count
                break
    
    # Handle π-stacking types
    if "π-Stacking" in interaction_types:
        if "parallel" in content.lower():
            interaction_types.add("π-Stacking (parallel)")
        if "perpendicular" in content.lower():
            interaction_types.add("π-Stacking (perpendicular)")
        interaction_types.discard("π-Stacking")
    
    return num_interactions, ", ".join(sorted(interaction_types))

def convert_plip_txt_to_xlsx(txt_path, xlsx_path):
    """
    Convert a PLIP TXT report to a formatted XLSX file.
    """
    sections = []
    current_section = None
    current_header = None
    current_data = []
    
    with open(txt_path, 'r') as f:
        lines = f.readlines()

    for line in lines:
        # Detect section headers
        match = re.match(r"\*\*(.+?)\*\*", line)
        if match:
            # Save previous section if exists
            if current_section:
                sections.append({
                    'type': current_section,
                    'header': current_header,
                    'data': current_data
                })
            current_section = match.group(1).strip()
            current_header = None
            current_data = []
            continue
        
        # Detect header line (first line with pipes after section header)
        if "|" in line and current_header is None:
            current_header = [item.strip() for item in line.strip().strip('|').split('|')]
            continue
            
        # Detect data lines (with pipes)
        if "|" in line and current_header is not None:
            # Remove leading/trailing pipes and split
            cleaned_line = [item.strip() for item in line.strip().strip('|').split('|')]
            current_data.append(cleaned_line)
    
    # Add the last section
    if current_section and current_header:
        sections.append({
            'type': current_section,
            'header': current_header,
            'data': current_data
        })
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Interactions"
    
    # Define styles
    orange_fill = PatternFill(start_color='EB641B', end_color='EB641B', fill_type='solid')
    blue_fill = PatternFill(start_color='92cddc', end_color='92cddc', fill_type='solid')
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 1
    
    for section in sections:
        section_start_row = current_row
        
        # Add section header
        ws.cell(row=current_row, column=1, value=section['type'])
        # Apply orange background and bold font to section header
        for col in range(1, len(section['header']) + 1):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = orange_fill
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        current_row += 1
        
        # Add column headers
        for col_idx, header in enumerate(section['header'], 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.fill = blue_fill
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        current_row += 1
        
        # Add data rows
        for data_row in section['data']:
            for col_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.alignment = center_alignment
                cell.border = thin_border
                
                # Format numbers as text to avoid Excel's warning
                if value and value.replace('.', '').replace(',', '').replace('-', '').isdigit():
                    cell.number_format = '@'  # Format as text
            
            current_row += 1
        
        # Apply borders to the entire section
        section_end_row = current_row - 1
        num_columns = len(section['header'])
        
        for row in range(section_start_row, section_end_row + 1):
            for col in range(1, num_columns + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
        
        # Add empty row after each section
        current_row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save Excel file
    wb.save(xlsx_path)

def get_interaction_data_from_xlsx(xlsx_path):
    """
    Parse XLSX file to extract interaction counts and types.
    Now correctly counts only interaction rows, skipping headers and empty rows.
    """
    try:
        import openpyxl
        
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        
        interaction_types = set()
        num_interactions = 0
        
        # Mapping from section names to standardized names
        section_mapping = {
            "HYDROPHOBIC INTERACTIONS": "Hydrophobic Interactions",
            "HYDROGEN BONDS": "Hydrogen Bonds",
            "WATER-BRIDGED HYDROGEN BONDS": "Water Bridges",
            "PI-STACKING": "π-Stacking",
            "PI-CATION INTERACTIONS": "π-Cation Interactions",
            "HALOGEN BONDS": "Halogen Bonds",
            "SALT BRIDGES": "Salt Bridges",
            "METAL COMPLEXES": "Metal Complexation"
        }
        
        current_section = None
        in_data_section = False
        
        for row in ws.iter_rows(values_only=True):
            if not any(row):
                continue
                
            # Check if this is a section header (first cell has value, others might be empty)
            if row[0] and any(keyword in str(row[0]).upper() for keyword in section_mapping.keys()):
                for keyword, std_name in section_mapping.items():
                    if keyword in str(row[0]).upper():
                        current_section = std_name
                        in_data_section = True
                        break
                continue
            
            # Skip the column header row that comes after section headers
            if in_data_section and row[0] and any(str(cell).upper() in ["RESNR", "RESTYPE", "RESCHAIN"] for cell in row if cell):
                in_data_section = False  # We've passed the header row
                continue
            
            # Check if this is a data row (has residue number in first column)
            if current_section and row[0] and str(row[0]).strip().isdigit():
                num_interactions += 1
                interaction_types.add(current_section)
        
        return num_interactions, ", ".join(sorted(interaction_types))
    
    except Exception as e:
        return 0, ""

def log_error(plip_output_dir, process, error_msg):
    """
    Log errors to the error log file in the plip_processed_files directory.
    """
    error_log_path = os.path.join(plip_output_dir, "error_log.txt")
    with open(error_log_path, 'a') as f:
        f.write(f"{process} error: {error_msg}\n")
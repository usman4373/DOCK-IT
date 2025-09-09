import os
import pandas as pd
import glob
import shutil
import streamlit as st
from pymol import cmd
from .pdb_utils import extract_protein_sequence
from .plip_processing import run_plip_analysis
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

def detect_smiles_columns(df):
    """Detect which columns in a DataFrame might contain SMILES strings"""
    possible_smiles_cols = []
    for col in df.columns:
        sample = df[col].dropna().head(10)
        if len(sample) > 0:
            smiles_pattern = r'[\[\]\(\)=#+-]'
            smiles_count = sample.astype(str).str.contains(smiles_pattern).sum()
            if smiles_count / len(sample) > 0.5:
                possible_smiles_cols.append(col)
    return possible_smiles_cols

def detect_id_columns(df, smiles_col):
    """Detect which columns in a DataFrame might contain ID values"""
    possible_id_cols = []
    for col in df.columns:
        if col != smiles_col:
            sample = df[col].dropna().head(10)
            if len(sample) > 0:
                smiles_pattern = r'[\[\]\(\)=#+-]'
                smiles_count = sample.astype(str).str.contains(smiles_pattern).sum()
                if smiles_count / len(sample) < 0.2:
                    possible_id_cols.append(col)
    return possible_id_cols

def map_proteins_to_reference_ligands(proteins_dir, reference_ligands_dir):
    """Map proteins to reference ligands based on sequence similarity"""
    protein_files = glob.glob(os.path.join(proteins_dir, "*.pdb"))
    ref_files = glob.glob(os.path.join(reference_ligands_dir, "*.pdb"))
    
    mapping = {}
    
    # Extract sequences from all reference files
    ref_sequences = {}
    for ref_file in ref_files:
        sequence = extract_protein_sequence(ref_file)
        if sequence:
            ref_sequences[ref_file] = sequence
    
    # Match each protein to a reference file
    for protein_file in protein_files:
        protein_sequence = extract_protein_sequence(protein_file)
        if not protein_sequence:
            st.session_state.error_log.append(f"Could not extract sequence from {protein_file}")
            continue
            
        # Find reference with matching sequence
        matched_ref = None
        for ref_file, ref_sequence in ref_sequences.items():
            if protein_sequence == ref_sequence:
                matched_ref = ref_file
                break
        
        if matched_ref:
            mapping[protein_file] = matched_ref
        else:
            st.session_state.error_log.append(f"No matching reference found for {protein_file}")
    
    return mapping

def process_docking_results(output_dir, minimize=False):
    """Process docking results with support for both minimized and multi-pose outputs"""
    try:
        # Find all protein directories
        protein_dirs = [d for d in os.listdir(output_dir) 
                       if os.path.isdir(os.path.join(output_dir, d)) and 
                       d not in ["log", "extracted_reference_ligands", "docking_results_per_protein"]]
        
        for protein_dir in protein_dirs:
            protein_path = os.path.join(output_dir, protein_dir)
            
            # Create organized subdirectories
            docked_files_dir = os.path.join(protein_path, "docked_files")
            complexes_dir = os.path.join(protein_path, "complexes")
            
            os.makedirs(docked_files_dir, exist_ok=True)
            os.makedirs(complexes_dir, exist_ok=True)
            
            # Only create splitted_models directory if minimization is not enabled
            if not minimize:
                splitted_models_dir = os.path.join(protein_path, "splitted_models")
                os.makedirs(splitted_models_dir, exist_ok=True)
            
            # Find all PDBQT files in the protein directory and move them to docked_files
            pdbqt_files = glob.glob(os.path.join(protein_path, "*.pdbqt"))
            for pdbqt_file in pdbqt_files:
                try:
                    # Move to docked_files directory
                    new_path = os.path.join(docked_files_dir, os.path.basename(pdbqt_file))
                    shutil.move(pdbqt_file, new_path)
                except Exception as e:
                    error_msg = f"Error moving file {pdbqt_file}: {str(e)}"
                    st.session_state.error_log.append(error_msg)
            
            # Now process files from the docked_files directory
            pdbqt_files = glob.glob(os.path.join(docked_files_dir, "*.pdbqt"))
            for pdbqt_file in pdbqt_files:
                try:
                    # Move to docked_files directory
                    new_path = os.path.join(docked_files_dir, os.path.basename(pdbqt_file))
                    shutil.move(pdbqt_file, new_path)
                except Exception as e:
                    error_msg = f"Error moving file {pdbqt_file}: {str(e)}"
                    st.session_state.error_log.append(error_msg)
            
            # Now process files from the docked_files directory
            pdbqt_files = glob.glob(os.path.join(docked_files_dir, "*.pdbqt"))
            
            for pdbqt_file in pdbqt_files:
                ligand_name = os.path.basename(pdbqt_file).replace('.pdbqt', '')
                ligand_name_only = ligand_name.replace(f"{protein_dir}_", "")
                
                # For minimized docking (single pose), directly create complex
                if minimize:
                    try:
                        # Load protein
                        protein_file = os.path.join(st.session_state.proteins_dir, f"{protein_dir}.pdb")
                        if not os.path.exists(protein_file):
                            # Try to find the protein file with different case or extension
                            protein_files = glob.glob(os.path.join(st.session_state.proteins_dir, f"{protein_dir}*"))
                            if protein_files:
                                protein_file = protein_files[0]
                        
                        # Initialize PyMOL
                        cmd.reinitialize()
                        
                        # Load protein and ligand
                        cmd.load(protein_file, "protein")
                        cmd.load(pdbqt_file, "ligand")
                        
                        # Create complex and save in complexes directory
                        complex_name = f"{protein_dir}-{ligand_name_only}-complex.pdb"
                        complex_path = os.path.join(complexes_dir, complex_name)
                        cmd.save(complex_path, "protein or ligand")
                        
                        # Get affinity from results
                        ligand_results = st.session_state.results_df[
                            (st.session_state.results_df['protein'] == protein_dir) & 
                            (st.session_state.results_df['ligand'] == ligand_name_only)
                        ]
                        
                        affinity = ligand_results.iloc[0]['affinity'] if not ligand_results.empty else "N/A"
                        
                        st.session_state.docking_log.append(
                            f"✓ Created complex for {protein_dir}-{ligand_name_only} "
                            f"(affinity: {affinity})"
                        )
                        
                        # Clean up
                        cmd.delete("all")
                        
                    except Exception as e:
                        error_msg = f"Error creating complex for {protein_dir}-{ligand_name}: {str(e)}"
                        st.session_state.error_log.append(error_msg)
                        continue
                
                # For multi-pose docking
                else:
                    # Create ligand-specific directory in splitted_models
                    ligand_models_dir = os.path.join(splitted_models_dir, ligand_name)
                    os.makedirs(ligand_models_dir, exist_ok=True)
                    
                    # Split multi-pose PDBQT file into individual poses
                    try:
                        # Initialize PyMOL
                        cmd.reinitialize()
                        
                        # Load the multi-pose PDBQT
                        cmd.load(pdbqt_file, "docked_ligand")
                        
                        # Check if there are multiple poses
                        num_states = cmd.count_states("docked_ligand")
                        
                        if num_states > 1:
                            # Split states into separate objects
                            cmd.split_states("docked_ligand")
                            
                            # Get all objects (poses)
                            objects = cmd.get_names("objects")
                            pose_objects = [obj for obj in objects if obj.startswith("docked_ligand_")]
                            
                            # Save each pose as separate PDB file in ligand-specific directory
                            for i, obj in enumerate(pose_objects, 1):
                                pose_name = f"{ligand_name}_pose_{i}"
                                pose_path = os.path.join(ligand_models_dir, f"{pose_name}.pdb")
                                cmd.save(pose_path, obj)
                            
                            st.session_state.docking_log.append(f"✓ Split {len(pose_objects)} poses for {ligand_name}")
                            
                        else:
                            # Only one pose, save it directly
                            pose_name = f"{ligand_name}_pose_1"
                            pose_path = os.path.join(ligand_models_dir, f"{pose_name}.pdb")
                            cmd.save(pose_path, "docked_ligand")
                            st.session_state.docking_log.append(f"✓ Saved single pose for {ligand_name}")
                        
                        # Clean up
                        cmd.delete("all")
                        
                    except Exception as e:
                        error_msg = f"Error processing poses for {ligand_name}: {str(e)}"
                        st.session_state.error_log.append(error_msg)
                        continue
                    
                    # Create protein-ligand complex for best pose
                    try:
                        # Find the best pose for this ligand from results
                        ligand_results = st.session_state.results_df[
                            (st.session_state.results_df['protein'] == protein_dir) & 
                            (st.session_state.results_df['ligand'] == ligand_name_only)
                        ]
                        
                        if not ligand_results.empty:
                            best_pose = int(ligand_results.iloc[0]['pose'])
                            best_affinity = ligand_results.iloc[0]['affinity']
                            
                            # Load protein
                            protein_file = os.path.join(st.session_state.proteins_dir, f"{protein_dir}.pdb")
                            if not os.path.exists(protein_file):
                                # Try to find the protein file with different case or extension
                                protein_files = glob.glob(os.path.join(st.session_state.proteins_dir, f"{protein_dir}*"))
                                if protein_files:
                                    protein_file = protein_files[0]
                            
                            cmd.load(protein_file, "protein")
                            
                            # Load the best pose
                            best_pose_path = os.path.join(ligand_models_dir, f"{ligand_name}_pose_{best_pose}.pdb")
                            cmd.load(best_pose_path, "best_pose")
                            
                            # Create complex and save in complexes directory
                            complex_name = f"{protein_dir}-{ligand_name_only}-complex.pdb"
                            complex_path = os.path.join(complexes_dir, complex_name)
                            cmd.save(complex_path, "protein or best_pose")
                            
                            st.session_state.docking_log.append(
                                f"✓ Created complex for {protein_dir}-{ligand_name_only} "
                                f"(affinity: {best_affinity:.2f}, pose: {best_pose})"
                            )
                            
                            # Clean up
                            cmd.delete("all")
                        
                    except Exception as e:
                        error_msg = f"Error creating complex for {protein_dir}-{ligand_name}: {str(e)}"
                        st.session_state.error_log.append(error_msg)
                        continue

        # Run PLIP analysis for each protein after all complexes are created
        for protein_dir in protein_dirs:
            protein_path = os.path.join(output_dir, protein_dir)
            complexes_dir = os.path.join(protein_path, "complexes")
            
            # Run PLIP analysis
            interaction_data = run_plip_analysis(complexes_dir, protein_path)
            
            # Update results with interaction data
            for complex_name, data in interaction_data.items():
                # Extract ligand name from complex name
                # Complex name format: {protein_dir}-{ligand_name_only}-complex
                if complex_name.startswith(protein_dir) and complex_name.endswith("-complex"):
                    ligand_name = complex_name.replace(f"{protein_dir}-", "").replace("-complex", "")
                    
                    # Update the results for this ligand
                    mask = (st.session_state.results_df['protein'] == protein_dir) & \
                           (st.session_state.results_df['ligand'] == ligand_name)
                    
                    if mask.any():
                        st.session_state.results_df.loc[mask, 'Number of Interactions'] = data['num_interactions']
                        st.session_state.results_df.loc[mask, 'Types of Interactions'] = data['interaction_types']
        
        # Save updated results with interaction data
        save_results_to_xlsx(output_dir)
        save_per_protein_results(output_dir)
        
        return True
        
    except Exception as e:
        st.session_state.error_log.append(f"Error in post-processing: {str(e)}")
        return False

def update_results_xlsx(protein_name, ligand_name, affinity, pose, output_dir, num_interactions=0, interaction_types=""):
    """
    Update results and save as formatted XLSX with interaction data.
    """
    # Check if this result already exists
    mask = (st.session_state.results_df['protein'] == protein_name) & \
           (st.session_state.results_df['ligand'] == ligand_name)
    
    if mask.any():
        # Update existing row
        st.session_state.results_df.loc[mask, 'affinity'] = affinity
        st.session_state.results_df.loc[mask, 'pose'] = pose
        # Don't update interaction data here - it will be updated later by PLIP
    else:
        # Add new row with default interaction values
        new_row = pd.DataFrame({
            "protein": [protein_name], 
            "ligand": [ligand_name], 
            "affinity": [affinity],
            "pose": [pose],
            "Number of Interactions": [num_interactions],
            "Types of Interactions": [interaction_types]
        })
        st.session_state.results_df = pd.concat([st.session_state.results_df, new_row], ignore_index=True)
    
    # Save to XLSX
    save_results_to_xlsx(output_dir)
    
    # Also save per-protein results
    save_per_protein_results(output_dir)

def save_results_to_xlsx(output_dir):
    """
    Save the results dataframe to a formatted XLSX file.
    """
    results_path = os.path.join(output_dir, "docking_results.xlsx")
    
    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Docking Results"
    
    # Define styles
    header_fill = PatternFill(start_color='EB641B', end_color='EB641B', fill_type='solid')
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Write headers
    for col_idx, column in enumerate(st.session_state.results_df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_idx, row in enumerate(st.session_state.results_df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_alignment
            cell.border = thin_border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(results_path)

def save_per_protein_results(output_dir):
    """
    Save per-protein results in a separate folder.
    """
    # Make sure we're not creating these folders inside the per_protein_dir
    per_protein_dir = os.path.join(output_dir, "docking_results_per_protein")
    os.makedirs(per_protein_dir, exist_ok=True)
    
    # Define styles
    header_fill = PatternFill(start_color='EB641B', end_color='EB641B', fill_type='solid')
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for protein in st.session_state.results_df['protein'].unique():
        protein_results = st.session_state.results_df[st.session_state.results_df['protein'] == protein]
        
        # Create a workbook for this protein
        wb = Workbook()
        ws = wb.active
        ws.title = "Docking Results"
        
        # Write headers
        for col_idx, column in enumerate(protein_results.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Write data rows
        for row_idx, row in enumerate(protein_results.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = center_alignment
                cell.border = thin_border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the file
        protein_file = os.path.join(per_protein_dir, f"{protein}.xlsx")
        wb.save(protein_file)
